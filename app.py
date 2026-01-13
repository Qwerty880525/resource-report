import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile

st.set_page_config(page_title="Отчет по ресурсам", layout="centered")
st.title("📊 Формирование отчета по ресурсам")

# ---------------- ЗАГРУЗКА ФАЙЛОВ ----------------

project_files = st.file_uploader(
    "1) Загрузите файлы проектов (Excel)",
    type=["xlsx"],
    accept_multiple_files=True
)

template_file = st.file_uploader(
    "2) Загрузите ШАБЛОН отчета (template.xlsx)",
    type=["xlsx"]
)

# ---------------- ПЕРИОД ----------------

col1, col2 = st.columns(2)
with col1:
    date_from = st.date_input("Начало периода")
with col2:
    date_to = st.date_input("Окончание периода")

generate = st.button("🚀 Сформировать отчет")

# ---------------- ФУНКЦИЯ ЧТЕНИЯ ----------------

def read_data(file):
    return pd.read_excel(file, sheet_name="Data")

# ---------------- ОСНОВНАЯ ЛОГИКА ----------------

if generate:

    if not project_files:
        st.error("Загрузите файлы проектов")
        st.stop()

    if not template_file:
        st.error("Загрузите шаблон отчета")
        st.stop()

    # 1. объединяем данные
    dfs = []
    for f in project_files:
        df = read_data(f)
        dfs.append(df)

    data = pd.concat(dfs, ignore_index=True)

    st.subheader("Выберите колонки дат")

    cols = data.columns.tolist()
    col_start = st.selectbox("Колонка НАЧАЛА", cols)
    col_end   = st.selectbox("Колонка ОКОНЧАНИЯ", cols)

    data[col_start] = pd.to_datetime(data[col_start], errors="coerce")
    data[col_end]   = pd.to_datetime(data[col_end], errors="coerce")

    # 2. фильтр по пересечению периодов
    mask = (
        (data[col_start] <= pd.to_datetime(date_to)) &
        (data[col_end]   >= pd.to_datetime(date_from))
    )

    filtered = data[mask]

    st.success(f"Найдено строк: {len(filtered)}")

    # 3. сохраняем шаблон во временный файл
    tmp_template = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_template.write(template_file.read())
    tmp_template.close()

    # 4. открываем шаблон
    wb = load_workbook(tmp_template.name)
    ws = wb["Data"]   # ЛИСТ В ШАБЛОНЕ

    # 5. чистим старые строки
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # 6. записываем новые данные
    for _, row in filtered.iterrows():
        ws.append(list(row))

    # 7. сохраняем готовый файл
    tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_out.name)

    with open(tmp_out.name, "rb") as f:
        st.download_button(
            "⬇ Скачать отчет",
            f,
            file_name="Отчет_по_ресурсам.xlsx"
        )
